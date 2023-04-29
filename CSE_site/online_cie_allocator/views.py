from django.shortcuts import render, redirect, HttpResponse
from django.urls import reverse
from .decorators import allowed_users
from django.core.mail import send_mail

from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from CSE_site.models import Course
from CSE_site.settings import EMAIL_HOST_USER
from authentication.models import Student, Faculty

def exam_allocator(req):
    if req.user.groups.filter(name="Office"):
        return redirect(reverse("exam_allocator_office_dashboard"))
    else:
        return HttpResponse("You don't have access to this page")

@allowed_users(["Office"])
def office_dashboard_view(req):
    return render(req, "exam_allocator_office_dashboard.html", {})

@allowed_users(["Office"])
def office_assign_students_view(req):
    if req.method == "POST":
        sem = int(req.POST.get("semester"))

        course_objects = [course_object for course_object in Course.objects.filter(sem=sem)]

        return render(req, "exam_allocator_office_assign_students.html", {"sem": sem,"course_objects": course_objects})
    return render(req, "exam_allocator_office_assign_students.html", {})

@allowed_users(["Office"])
def office_assign_students_download_template(req, course_code):
    wb = Workbook()

    # TODO: make this course_code instead of short_code (code) used here
    course_object = Course.objects.get(code=course_code)
    wb.create_sheet(course_object.code)
    ws = wb[course_object.code]
    ws.append(["USN", "Room"])
    for i, student_object in enumerate(Student.objects.filter(courses=course_object)):
            ws[f"A{i+2}"] = student_object.usn          

    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    excel_data = save_virtual_workbook(wb)
    response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # TODO: make this course_code instead of short_code (code) used here
    response['Content-Disposition'] = f'attachment; filename="{course_code}_Exam_Room_Allocation.xlsx"'

    return response

@allowed_users(["Office"])
def office_assign_students_upload(req):
    if req.method == "POST":
        exam_title = req.POST.get("exam")
        try:
            excel_file = req.FILES["excel_file"]
        except Exception as e:
            return HttpResponse(f"Please select a file before uploading {e}")

        if not excel_file.name.endswith("xlsx"):
            return HttpResponse("Please upload a .xlsx file")
        
        wb = load_workbook(excel_file)
        ws = wb.active

        email_subject = f"Assigned room for {ws.title} - {exam_title}"

        rows_iter = ws.iter_rows()
        next(rows_iter)
        for row in rows_iter:
            student = Student.objects.get(usn=row[0].value)
            email_body = f"Dear {student.name}, you have to assigned to room {row[1].value} for your {ws.title} test in {exam_title}"
            send_mail(email_subject, email_body, EMAIL_HOST_USER, [student.email])
        return HttpResponse("Email has been sent")

@allowed_users(["Office"])
def office_assign_invigilators_view(req):
    return render(req, "exam_allocator_office_assign_invigilators.html", {})

@allowed_users(["Office"])
def office_assign_invigilators_download_template(req):
    wb = Workbook()
    ws = wb.active

    ws.append(["Name", "Room"])
    for i, faculty_object in enumerate(Faculty.objects.all()):
            ws[f"A{i+2}"] = faculty_object.name

    excel_data = save_virtual_workbook(wb)
    response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Invigilators_Exam_Room_Allocation.xlsx"'

    return response

@allowed_users(["Office"])
def office_assign_invigilators_upload(req):
    if req.method == "POST":
        date = req.POST.get("date")
        time = req.POST.get("time")

        try:
            excel_file = req.FILES["excel_file"]
        except Exception as e:
            return HttpResponse(f"Please select a file before uploading {e}")

        if not excel_file.name.endswith("xlsx"):
            return HttpResponse("Please upload a .xlsx file")
        
        wb = load_workbook(excel_file)
        ws = wb.active

        email_subject = f"Assigned room for Invigilation"
        rows_iter = ws.iter_rows()
        next(rows_iter)
        for row in rows_iter:
            # TODO: make sure to make a unique primary key for faculties
            faculty = Faculty.objects.get(name=row[0].value)
            email_body = f"Dear {faculty.name}, you have to assigned to room {row[1].value} as invigilator on {date} at {time}"
            send_mail(email_subject, email_body, EMAIL_HOST_USER, [faculty.email])
        return HttpResponse("Email has been sent")