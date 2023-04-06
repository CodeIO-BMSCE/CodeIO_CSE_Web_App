from django.shortcuts import render, redirect, HttpResponse, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .models import Office
from student_dashboard_proctor.models import Student, StudentDetail, Sem, Fastrack
from faculty_dashboard_proctor.models import Faculty
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
from django.conf import settings
import openpyxl
import pandas as pd


@login_required
def dashboard(request):
    office = Office.objects.get(email=request.user.email)
    faculty = Faculty.objects.all()
    context = {'office': office, 'faculties': faculty}
    return render(request, 'office_dashboard_proctor/dashboard.html', context)


def proctor_assign(request):
    if request.method == "POST":
        excel_file = request.FILES['excel']
        df = pd.read_excel(excel_file)
        for index, row in df.iterrows():
            std_usn = row['USN']
            proctor_email = row['Proctor-Email-ID']
            try:
                student = Student.objects.get(USN=std_usn)
                proctor = Faculty.objects.get(email=proctor_email)
                student.proctor_id = proctor
                student.save()
            except (Student.DoesNotExist, Faculty.DoesNotExist):
                pass
        return redirect('office:dashboard')
    return render(request, 'office_dashboard_proctor/proctor_assign.html')



def download_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Sl. No.'
    ws['B1'] = 'USN'
    ws['C1'] = 'Proctor-Name'
    ws['D1'] = 'Proctor-Email-ID'
    filename = 'proctor_assignment_sheet.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def view_students(request,pk):
    faculty=Faculty.objects.get(email=pk)
    students=Student.objects.filter(proctor_id=faculty)
    context = {'students':students,'faculty':faculty}
    return render(request,'office_dashboard_proctor/view_students.html',context)

# def remove_student(request,pk):